"""
GPS vs MIS Fleet Dashboard - Streamlit version
================================================
Same dashboard as the web app, but as a Streamlit app you run on your own
machine or company server. Gives you a URL that works on your company
network (Wi-Fi / LAN) without ever touching the public internet.

RUN:
    pip install -r requirements.txt
    streamlit run app.py

Streamlit will print two URLs:
    Local URL:   http://localhost:8501       (only this PC)
    Network URL: http://192.168.x.x:8501     (anyone on your company network)

Share the Network URL with your team - it only works for people on the same
office Wi-Fi / LAN, never reaches the public internet.
"""

import io
import json
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(page_title="GPS vs MIS Fleet Dashboard", layout="wide", page_icon="🚚", initial_sidebar_state="expanded")

# ---------------------------------------------------------------------------
# Theme (light, clean SaaS-dashboard style — white cards, colored top borders)
# ---------------------------------------------------------------------------
COLOR_BG = "#F4F6F9"
COLOR_PANEL = "#FFFFFF"
COLOR_BORDER = "#E5E8EC"
COLOR_TEXT = "#1F2733"
COLOR_MUTED = "#6B7684"
COLOR_GPS = "#0EA5E9"
COLOR_MIS = "#F59E0B"
COLOR_OK = "#10B981"
COLOR_CRIT = "#EF4444"
COLOR_WARN = "#F59E0B"
COLOR_DIFF = "#8B5CF6"
COLOR_VEHICLES = "#3B82F6"

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@500;600;700&family=Inter:wght@400;500;600;700&family=IBM+Plex+Mono:wght@400;500&display=swap');

html, body, [class*="css"] {{ font-family: 'Inter', sans-serif; font-size: 13px; }}
.stApp {{ background-color: {COLOR_BG}; color: {COLOR_TEXT}; }}
#MainMenu {{ visibility: hidden; }}
footer {{ visibility: hidden; }}

h1 {{ font-family: 'Space Grotesk', sans-serif !important; font-weight: 700 !important; color: {COLOR_TEXT} !important; font-size: 26px !important; }}
h2, h3 {{ font-family: 'Space Grotesk', sans-serif !important; font-weight: 600 !important; color: {COLOR_TEXT} !important; font-size: 17px !important; }}
h4 {{ font-family: 'Space Grotesk', sans-serif !important; font-weight: 600 !important; color: {COLOR_TEXT} !important; font-size: 14px !important; }}
p, label, span, div {{ font-size: 13px; }}

/* Custom KPI cards */
.kpi-card {{
    background: {COLOR_PANEL}; border: 1px solid {COLOR_BORDER}; border-radius: 12px;
    padding: 18px 18px 16px; position: relative; overflow: hidden;
    box-shadow: 0 1px 3px rgba(16,24,40,0.06); transition: transform 0.15s ease, box-shadow 0.15s ease;
}}
.kpi-card:hover {{ transform: translateY(-2px); box-shadow: 0 6px 16px rgba(16,24,40,0.10); }}
.kpi-topbar {{ position: absolute; top: 0; left: 0; right: 0; height: 4px; }}
.kpi-icon {{
    width: 38px; height: 38px; border-radius: 10px; display: flex; align-items: center;
    justify-content: center; font-size: 18px; margin-bottom: 14px;
}}
.kpi-label {{ color: {COLOR_MUTED}; font-size: 11px; letter-spacing: 0.6px; text-transform: uppercase; font-weight: 600; margin-bottom: 4px; }}
.kpi-value {{ color: {COLOR_TEXT}; font-family: 'IBM Plex Mono', monospace; font-size: 24px; font-weight: 600; margin-bottom: 8px; }}
.kpi-delta {{ display: inline-flex; align-items: center; gap: 4px; font-size: 11px; font-weight: 600; padding: 3px 9px; border-radius: 20px; font-family: 'IBM Plex Mono', monospace; }}

.action-card {{
    background: {COLOR_PANEL}; border: 1px solid {COLOR_BORDER}; border-radius: 8px;
    padding: 10px 14px; margin-bottom: 6px;
}}

/* Dataframe / tables */
.stDataFrame, [data-testid="stDataFrame"] {{
    border: 1px solid {COLOR_BORDER} !important; border-radius: 10px !important; overflow: hidden;
}}
[data-testid="stDataFrame"] * {{ font-family: 'IBM Plex Mono', monospace !important; font-size: 11px !important; }}

/* Inputs */
.stTextInput input, .stSelectbox [data-baseweb="select"] > div, .stTextInput > div > div {{
    background: {COLOR_PANEL} !important; border: 1px solid {COLOR_BORDER} !important;
    color: {COLOR_TEXT} !important; border-radius: 7px !important; font-size: 13px !important;
}}
.stSlider [data-baseweb="slider"] {{ padding-top: 6px; }}
.stSlider [role="slider"] {{ background: {COLOR_MIS} !important; box-shadow: 0 0 0 4px rgba(245,158,11,0.18) !important; }}
.stSlider div[data-baseweb="slider"] > div > div {{ background: linear-gradient(90deg, {COLOR_GPS}, {COLOR_MIS}) !important; }}

/* File uploader */
[data-testid="stFileUploaderDropzone"] {{
    background: linear-gradient(135deg, rgba(14,165,233,0.04), {COLOR_PANEL} 70%) !important;
    border: 1.5px dashed {COLOR_BORDER} !important; border-radius: 12px !important;
    transition: border-color 0.15s ease;
}}
[data-testid="stFileUploaderDropzone"]:hover {{ border-color: {COLOR_GPS} !important; }}

/* Buttons — normal size by default */
.stButton button, .stFormSubmitButton button {{
    background: {COLOR_GPS} !important; color: #FFFFFF !important; border: none !important;
    border-radius: 7px !important; font-weight: 600 !important; font-size: 12px !important;
    padding: 0.35rem 0.8rem !important; transition: transform 0.1s ease, opacity 0.1s ease;
    box-shadow: 0 1px 3px rgba(14,165,233,0.3);
}}
.stButton button:hover, .stFormSubmitButton button:hover {{ opacity: 0.88; transform: translateY(-1px); }}

/* Corrective-action buttons only — bigger, card-like, two-line label */
.st-key-corrective_actions button {{
    white-space: pre-line !important; line-height: 1.3 !important; min-height: 54px !important;
    font-family: 'IBM Plex Mono', monospace !important; font-size: 11px !important; font-weight: 700 !important;
    box-shadow: 0 2px 8px rgba(16,24,40,0.12);
}}

/* Corrective-action buttons colored by severity */
[class*="st-key-actsev_critical"] button {{
    background: linear-gradient(135deg, #F87171, #EF4444) !important; color: #FFF !important;
}}
[class*="st-key-actsev_warn"] button {{
    background: linear-gradient(135deg, #FBBF24, #F59E0B) !important; color: #4A2E00 !important;
}}
[class*="st-key-actsev_watch"] button {{
    background: linear-gradient(135deg, #FCD34D, #F59E0B) !important; color: #4A2E00 !important; opacity: 0.92;
}}
[class*="st-key-actsev_ok"] button {{
    background: linear-gradient(135deg, #34D399, #10B981) !important; color: #FFF !important;
}}

/* Sidebar */
[data-testid="stSidebar"] {{ background: {COLOR_PANEL} !important; border-right: 1px solid {COLOR_BORDER}; }}
[data-testid="stSidebar"] img {{ border-radius: 6px; }}

/* Captions */
.stCaption, [data-testid="stCaptionContainer"] {{ color: {COLOR_MUTED} !important; font-family: 'IBM Plex Mono', monospace !important; }}

/* Checkbox label */
.stCheckbox label p {{ color: {COLOR_TEXT} !important; font-size: 13px !important; }}

/* Alert boxes */
[data-testid="stAlert"] {{ background: {COLOR_PANEL} !important; border: 1px solid {COLOR_BORDER} !important; border-radius: 8px; }}
</style>
""", unsafe_allow_html=True)

ACTION_COLORS = {"critical": COLOR_CRIT, "warn": COLOR_WARN, "watch": COLOR_MIS, "ok": COLOR_OK}


def kpi_card(container, icon, accent, label, value, delta_text=None, delta_positive=None, neutral=False):
    if delta_text and neutral:
        delta_html = f"<span class='kpi-delta' style='background:{COLOR_BORDER}; color:{COLOR_MUTED};'>{delta_text}</span>"
    elif delta_text:
        d_bg = "rgba(16,185,129,0.12)" if delta_positive else "rgba(239,68,68,0.12)"
        d_color = COLOR_OK if delta_positive else COLOR_CRIT
        arrow = "▲" if delta_positive else "▼"
        delta_html = f"<span class='kpi-delta' style='background:{d_bg}; color:{d_color};'>{arrow} {delta_text}</span>"
    else:
        delta_html = ""
    container.markdown(
        f"<div class='kpi-card'>"
        f"<div class='kpi-topbar' style='background:{accent};'></div>"
        f"<div class='kpi-icon' style='background:{accent}22; color:{accent};'>{icon}</div>"
        f"<div class='kpi-label'>{label}</div>"
        f"<div class='kpi-value'>{value}</div>"
        f"{delta_html}"
        f"</div>",
        unsafe_allow_html=True,
    )



# ---------------------------------------------------------------------------
# Login credentials + company branding (stored on disk, editable at runtime)
# ---------------------------------------------------------------------------
CREDENTIALS_PATH = Path(__file__).parent / ".streamlit" / "credentials.json"
BRANDING_PATH = Path(__file__).parent / ".streamlit" / "branding.json"
LOGO_PATH = Path(__file__).parent / ".streamlit" / "logo.png"


def load_credentials():
    # On Streamlit Community Cloud, prefer the Secrets manager (configured in the
    # app's dashboard, never stored in the git repo) over the local file.
    try:
        if "ADMIN_USERNAME" in st.secrets and "ADMIN_PASSWORD" in st.secrets:
            return {"username": st.secrets["ADMIN_USERNAME"], "password": st.secrets["ADMIN_PASSWORD"]}
    except Exception:
        pass
    if CREDENTIALS_PATH.exists():
        try:
            return json.loads(CREDENTIALS_PATH.read_text())
        except Exception:
            return None
    return None


def using_cloud_secrets():
    try:
        return "ADMIN_USERNAME" in st.secrets and "ADMIN_PASSWORD" in st.secrets
    except Exception:
        return False


def save_credentials(creds):
    CREDENTIALS_PATH.parent.mkdir(exist_ok=True)
    CREDENTIALS_PATH.write_text(json.dumps(creds, indent=2))


def load_branding():
    if BRANDING_PATH.exists():
        try:
            return json.loads(BRANDING_PATH.read_text())
        except Exception:
            return {}
    return {}


def save_branding(data):
    BRANDING_PATH.parent.mkdir(exist_ok=True)
    BRANDING_PATH.write_text(json.dumps(data, indent=2))


# ---------------------------------------------------------------------------
# Login
# ---------------------------------------------------------------------------
def check_login():
    if st.session_state.get("authenticated"):
        return True

    st.markdown(
        "<div style='color:#0EA5E9; font-family:\"IBM Plex Mono\",monospace; font-size:12px; letter-spacing:2px;'>FLEET TELEMETRY RECONCILIATION</div>"
        "<h1 style='margin-top:2px;'>GPS vs MIS Dashboard</h1>",
        unsafe_allow_html=True,
    )
    st.markdown("#### Admin login")
    with st.form("login_form"):
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Log in")

    if submitted:
        creds = load_credentials()
        if not creds:
            st.error(
                "No admin credentials configured yet. Run start.bat again, or copy "
                "`.streamlit/credentials.example.json` to `.streamlit/credentials.json` "
                "and set your own username/password (see README.md)."
            )
            return False
        if username == creds.get("username") and password == creds.get("password"):
            st.session_state.authenticated = True
            st.session_state.current_username = username
            st.rerun()
        else:
            st.error("Incorrect username or password.")
    return False


if not check_login():
    st.stop()

# Permanent company branding — edit this constant directly to change the
# name. To change the logo, replace the file at .streamlit/logo.png.
COMPANY_NAME = "Supreme Facility Management Limited"

with st.sidebar:
    if LOGO_PATH.exists():
        st.image(str(LOGO_PATH), width=100)
    st.markdown(f"### {COMPANY_NAME}")

    st.caption(f"Logged in as **{st.session_state.get('current_username', 'admin')}**")
    if st.button("Log out"):
        st.session_state.authenticated = False
        st.rerun()

    with st.expander("🔑 Change password"):
        if using_cloud_secrets():
            st.info(
                "This app is running on Streamlit Community Cloud, where the password "
                "comes from the app's **Secrets** settings, not this file. To change it: "
                "go to your app on share.streamlit.io → Settings → Secrets, update "
                "ADMIN_PASSWORD there, and save (the app restarts automatically)."
            )
        else:
            with st.form("change_password_form"):
                old_pw = st.text_input("Current password", type="password")
                new_pw = st.text_input("New password", type="password")
                confirm_pw = st.text_input("Confirm new password", type="password")
                pw_submit = st.form_submit_button("Update password")
            if pw_submit:
                creds = load_credentials() or {}
                if old_pw != creds.get("password"):
                    st.error("Current password is incorrect.")
                elif not new_pw:
                    st.error("New password can't be empty.")
                elif new_pw != confirm_pw:
                    st.error("New password and confirmation don't match.")
                else:
                    creds["password"] = new_pw
                    save_credentials(creds)
                    st.success("Password updated — use it next time you log in.")

# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------
def read_sheet(ws):
    day_cols = []
    for c in range(4, ws.max_column + 1):
        h = ws.cell(row=4, column=c).value
        if isinstance(h, (int, float)):
            day_cols.append(c)
        elif h is None and day_cols:
            break
    rows = {}
    for r in range(5, ws.max_row + 1):
        veh = ws.cell(row=r, column=2).value
        if not veh:
            continue
        cc = ws.cell(row=r, column=3).value
        vals = []
        for c in day_cols:
            v = ws.cell(row=r, column=c).value
            vals.append(v if isinstance(v, (int, float)) else 0)
        rows[str(veh).strip()] = {"cost_center": (cc or "Unassigned"), "days": vals}
    return rows


@st.cache_data(show_spinner="Reading workbook...")
def build_dataset(file_bytes):
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    gps_name = next((n for n in wb.sheetnames if n.strip().upper() == "GPS"), None)
    mis_name = next((n for n in wb.sheetnames if n.strip().upper() == "MIS"), None)
    if not gps_name or not mis_name:
        raise ValueError('Could not find sheets named "GPS" and "MIS" in this file.')

    gps_rows = read_sheet(wb[gps_name])
    mis_rows = read_sheet(wb[mis_name])
    n_days = max(
        max((len(v["days"]) for v in gps_rows.values()), default=0),
        max((len(v["days"]) for v in mis_rows.values()), default=0),
    )
    all_veh = list(dict.fromkeys(list(gps_rows.keys()) + list(mis_rows.keys())))

    records = []
    daily_records = []
    for veh in all_veh:
        g = gps_rows.get(veh)
        m = mis_rows.get(veh)
        cc = (g or m)["cost_center"]
        gdays = (g["days"] if g else []) + [0] * n_days
        mdays = (m["days"] if m else []) + [0] * n_days
        gdays, mdays = gdays[:n_days], mdays[:n_days]
        tg, tm = sum(gdays), sum(mdays)
        td = tm - tg
        dpct = (td / tg) if tg else 0.0
        source = "Both" if (g and m) else ("GPS only" if g else "MIS only")
        records.append({
            "Vehicle": veh, "Site": cc, "Total GPS": round(tg, 1), "Total MIS": round(tm, 1),
            "Diff": round(td, 1), "Diff %": dpct, "Source": source,
        })
        for d in range(n_days):
            daily_records.append({"Vehicle": veh, "Day": d + 1, "GPS": gdays[d], "MIS": mdays[d]})

    return pd.DataFrame(records), pd.DataFrame(daily_records), n_days


def classify(row, threshold_pct):
    t = threshold_pct / 100
    tg, tm, dpct = row["Total GPS"], row["Total MIS"], row["Diff %"]
    if tg == 0 and tm > 0:
        return "Check GPS device", "critical", "No GPS km recorded all month while MIS shows movement. Confirm the device is powered, fitted, and reporting."
    if tm == 0 and tg > 0:
        return "File missing MIS log", "critical", "GPS shows movement but no MIS entries were filed. Follow up with the site log-keeper."
    if tg == 0 and tm == 0:
        return "No data either side", "warn", "Both sources show zero for the month. Confirm the vehicle was actually in service."
    if abs(dpct) > t:
        if dpct > 0:
            return "Audit MIS entries", "critical", f"MIS log shows {abs(dpct)*100:.1f}% more km than GPS. Cross-check trip sheets and fuel/DA claims for over-reporting."
        return "Verify unrecorded trips", "critical", f"GPS shows {abs(dpct)*100:.1f}% more km than MIS. Check for trips run but not logged."
    if abs(dpct) > t * 0.5:
        return "Keep an eye on it", "watch", f"Diff is under the {threshold_pct}% threshold but trending up. Worth a spot-check next month."
    return "No action needed", "ok", "GPS and MIS agree within tolerance."


# ---------------------------------------------------------------------------
# Month-over-month history (saved to disk, survives restarts, never
# overwritten by a new upload unless it's the same month being re-saved)
# ---------------------------------------------------------------------------
HISTORY_PATH = Path(__file__).parent / "history.json"

MONTH_MAP = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def load_history():
    if HISTORY_PATH.exists():
        try:
            return json.loads(HISTORY_PATH.read_text())
        except Exception:
            return {}
    return {}


def save_history(hist):
    HISTORY_PATH.write_text(json.dumps(hist, indent=2))


def guess_month_from_filename(filename):
    m = re.search(r"(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\D{0,5}(20\d{2})", filename, re.IGNORECASE)
    if m:
        mon = MONTH_MAP[m.group(1).lower()]
        year = int(m.group(2))
        key = f"{year:04d}-{mon:02d}"
        label = datetime(year, mon, 1).strftime("%B %Y")
        return key, label
    now = datetime.now()
    return now.strftime("%Y-%m"), now.strftime("%B %Y")


def get_previous_entry(hist, current_key):
    keys = sorted(k for k in hist.keys() if k < current_key)
    return hist[keys[-1]] if keys else None


# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------
st.markdown(
    f"<div style='color:{COLOR_GPS}; font-family:\"IBM Plex Mono\",monospace; font-size:12px; letter-spacing:2px;'>FLEET TELEMETRY RECONCILIATION</div>"
    f"<h1 style='margin-top:2px; background:linear-gradient(90deg,{COLOR_GPS},{COLOR_MIS}) !important; "
    f"-webkit-background-clip:text !important; -webkit-text-fill-color:transparent !important; background-clip:text !important; "
    f"display:inline-block;'>GPS vs MIS Dashboard</h1>",
    unsafe_allow_html=True,
)

uploaded = st.file_uploader("Drop this month's file here (needs GPS and MIS sheets)", type=["xlsx", "xls"])
hist = load_history()
sorted_hist_entries = sorted(hist.values(), key=lambda x: x["key"], reverse=True)
month_options = [e["label"] for e in sorted_hist_entries]

if uploaded is None and not month_options:
    st.info("Upload your monthly GPS + MIS workbook to see the analysis.")
    st.stop()

view_options = (["📤 Uploaded file"] if uploaded is not None else []) + month_options
view_choice = st.selectbox("Viewing", view_options, index=0)

has_daily = False

if view_choice == "📤 Uploaded file":
    try:
        df, daily_df, n_days = build_dataset(uploaded.getvalue())
    except Exception as e:
        st.error(str(e))
        st.stop()
    has_daily = True

    guessed_key, guessed_label = guess_month_from_filename(uploaded.name)
    month_label = st.text_input("Month label for this file (edit if the guess is wrong)", value=guessed_label)
    try:
        month_key = datetime.strptime(month_label.strip(), "%B %Y").strftime("%Y-%m")
    except ValueError:
        month_key = guessed_key
else:
    entry = hist[next(e["key"] for e in sorted_hist_entries if e["label"] == view_choice)]
    df = pd.DataFrame(entry["vehicle_totals"])
    df["Diff"] = df["Total MIS"] - df["Total GPS"]
    df["Diff %"] = df.apply(lambda r: (r["Diff"] / r["Total GPS"]) if r["Total GPS"] else 0, axis=1)
    df["Source"] = "Both"
    month_key = entry["key"]

    daily_compact = entry.get("daily")
    if daily_compact:
        has_daily = True
        daily_rows = []
        for veh, series in daily_compact.items():
            gvals, mvals = series["g"], series["m"]
            for d in range(len(gvals)):
                daily_rows.append({"Vehicle": veh, "Day": d + 1, "GPS": gvals[d], "MIS": mvals[d]})
        daily_df = pd.DataFrame(daily_rows)
        st.caption(f"Viewing saved history for **{view_choice}** — including day-wise breakdown.")
    else:
        daily_df = pd.DataFrame(columns=["Vehicle", "Day", "GPS", "MIS"])
        st.caption(f"Viewing saved history for **{view_choice}** — this month was saved before day-wise data was stored, so only totals are available.")

threshold = st.slider("Flag threshold (%)  — flag vehicles where |Diff %| exceeds this", 5, 100, 20, step=5)

action_info = df.apply(lambda r: classify(r, threshold), axis=1)
df["Action"] = action_info.apply(lambda x: x[0])
df["Severity"] = action_info.apply(lambda x: x[1])
df["ActionDetail"] = action_info.apply(lambda x: x[2])

total_gps = df["Total GPS"].sum()
total_mis = df["Total MIS"].sum()
overall_diff = total_mis - total_gps
overall_pct = (overall_diff / total_gps) if total_gps else 0
flagged = (df["Action"] != "No action needed").sum()

if view_choice == "📤 Uploaded file":
    daily_compact = {}
    for veh, grp in daily_df.groupby("Vehicle"):
        grp_sorted = grp.sort_values("Day")
        daily_compact[veh] = {
            "g": [round(float(v), 1) for v in grp_sorted["GPS"].tolist()],
            "m": [round(float(v), 1) for v in grp_sorted["MIS"].tolist()],
        }
    hist[month_key] = {
        "key": month_key, "label": month_label.strip(), "saved_at": datetime.now().isoformat(),
        "total_gps": float(total_gps), "total_mis": float(total_mis),
        "vehicles": int(len(df)), "sites": int(df["Site"].nunique()), "flagged": int(flagged),
        "site_summary": df.groupby("Site", as_index=False).agg(
            Total_GPS=("Total GPS", "sum"), Total_MIS=("Total MIS", "sum")
        ).to_dict("records"),
        "vehicle_totals": df[["Vehicle", "Site", "Total GPS", "Total MIS"]].to_dict("records"),
        "daily": daily_compact,
    }
    save_history(hist)
    st.caption(f"✓ Autosaved as **{month_label.strip()}** — {len(hist)} month(s) in history now.")

prev_entry = get_previous_entry(hist, month_key)
if prev_entry:
    gps_delta = total_gps - prev_entry["total_gps"]
    mis_delta = total_mis - prev_entry["total_mis"]
    gps_delta_label = f"{gps_delta:+,.0f} km vs {prev_entry['label']}"
    mis_delta_label = f"{mis_delta:+,.0f} km vs {prev_entry['label']}"
else:
    gps_delta_label = mis_delta_label = None
    st.caption("No previous month saved yet to compare against — once you save another month, comparisons will show here automatically.")

c1, c2, c3, c4, c5 = st.columns(5)
gps_positive = prev_entry is not None and (total_gps - prev_entry["total_gps"]) >= 0
mis_positive = prev_entry is not None and (total_mis - prev_entry["total_mis"]) >= 0
kpi_card(c1, "📡", COLOR_GPS, "TOTAL GPS KM", f"{total_gps:,.0f}",
         gps_delta_label.replace(" km", "") if gps_delta_label else None, gps_positive)
kpi_card(c2, "📝", COLOR_MIS, "TOTAL MIS KM", f"{total_mis:,.0f}",
         mis_delta_label.replace(" km", "") if mis_delta_label else None, mis_positive)
kpi_card(c3, "⚠️", COLOR_DIFF, "OVERALL DIFF", f"{overall_pct*100:.1f}%",
         f"{overall_diff:+,.0f} km", overall_diff >= 0)
kpi_card(c4, "🚚", COLOR_VEHICLES, "VEHICLES", f"{len(df)}",
         f"{df['Site'].nunique()} sites", neutral=True)
kpi_card(c5, "🚩", COLOR_CRIT, "FLAGGED", f"{flagged}",
         f"{flagged/len(df)*100:.0f}% of fleet", neutral=True)

if has_daily:
    st.markdown("### Day-wise total km, all vehicles")
    trend = daily_df.groupby("Day", as_index=False)[["GPS", "MIS"]].sum()
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=trend["Day"], y=trend["GPS"], name="GPS", mode="lines+markers",
        line=dict(color=COLOR_GPS, width=2.5), marker=dict(size=5),
        hovertemplate="Day %{x}<br>GPS: %{y:,.0f} km<extra></extra>",
    ))
    fig.add_trace(go.Scatter(
        x=trend["Day"], y=trend["MIS"], name="MIS", mode="lines+markers",
        line=dict(color=COLOR_MIS, width=2.5), marker=dict(size=5),
        hovertemplate="Day %{x}<br>MIS: %{y:,.0f} km<extra></extra>",
    ))
    fig.update_layout(
        plot_bgcolor=COLOR_PANEL, paper_bgcolor=COLOR_BG, font_color=COLOR_TEXT,
        height=300, margin=dict(l=10, r=10, t=10, b=10),
        xaxis=dict(gridcolor=COLOR_BORDER, title="Day of month", tickfont=dict(size=11), dtick=2),
        yaxis=dict(gridcolor=COLOR_BORDER, title="Km", tickfont=dict(size=11)),
        legend=dict(orientation="h", y=1.12, font=dict(size=12)),
        hovermode="x unified",
    )
    st.plotly_chart(fig, use_container_width=True)

st.markdown("### Sites")
site_summary = df.groupby("Site", as_index=False).agg(
    Vehicles=("Vehicle", "count"), Total_GPS=("Total GPS", "sum"), Total_MIS=("Total MIS", "sum"),
    Flagged=("Action", lambda s: (s != "No action needed").sum()),
)
site_summary["Diff %"] = ((site_summary["Total_MIS"] - site_summary["Total_GPS"]) / site_summary["Total_GPS"].replace(0, pd.NA)).fillna(0)
site_summary = site_summary.sort_values("Diff %", key=abs, ascending=False)


def site_severity(diff_pct, threshold_pct):
    t = threshold_pct / 100
    if abs(diff_pct) > t:
        return "critical"
    if abs(diff_pct) > t * 0.5:
        return "watch"
    return "ok"


site_summary["Severity"] = site_summary["Diff %"].apply(lambda d: site_severity(d, threshold))
site_bar_colors = site_summary["Severity"].map(ACTION_COLORS)

n_sites = len(site_summary)
bar_height = max(320, n_sites * 30)

# --- Chart 1: GPS vs MIS volume by site (split into two columns to keep it compact) ---
def make_site_volume_chart(data, chart_height):
    fig = go.Figure()
    fig.add_trace(go.Bar(
        y=data["Site"], x=data["Total_GPS"], name="GPS", orientation="h",
        marker_color=COLOR_GPS, text=data["Total_GPS"].apply(lambda v: f"{v:,.0f}"),
        textposition="outside", textfont=dict(size=10, color=COLOR_TEXT, family="IBM Plex Mono"),
        hovertemplate="%{y}<br>GPS: %{x:,.0f} km<extra></extra>",
    ))
    fig.add_trace(go.Bar(
        y=data["Site"], x=data["Total_MIS"], name="MIS", orientation="h",
        marker_color=COLOR_MIS, text=data["Total_MIS"].apply(lambda v: f"{v:,.0f}"),
        textposition="outside", textfont=dict(size=10, color=COLOR_TEXT, family="IBM Plex Mono"),
        hovertemplate="%{y}<br>MIS: %{x:,.0f} km<extra></extra>",
    ))
    fig.update_layout(
        barmode="group", plot_bgcolor=COLOR_PANEL, paper_bgcolor=COLOR_BG, font_color=COLOR_TEXT,
        height=chart_height, margin=dict(l=10, r=50, t=10, b=30),
        xaxis=dict(gridcolor=COLOR_BORDER, title="Km", tickfont=dict(size=10)),
        yaxis=dict(tickfont=dict(size=11, family="Inter"), automargin=True),
        legend=dict(orientation="h", y=1.05, x=0, font=dict(size=11)),
        bargap=0.28, bargroupgap=0.08,
    )
    return fig


vol_desc = site_summary.sort_values("Total_GPS", ascending=False).reset_index(drop=True)
half = (n_sites + 1) // 2
top_half = vol_desc.iloc[:half].sort_values("Total_GPS", ascending=True)
bottom_half = vol_desc.iloc[half:].sort_values("Total_GPS", ascending=True)
half_height = max(280, max(len(top_half), len(bottom_half)) * 30)

st.caption("GPS vs MIS total km by site — sorted by GPS volume, split into two for readability")
col_v1, col_v2 = st.columns(2)
with col_v1:
    st.plotly_chart(make_site_volume_chart(top_half, half_height), use_container_width=True)
with col_v2:
    st.plotly_chart(make_site_volume_chart(bottom_half, half_height), use_container_width=True)

# --- Chart 2: Diff % by site (vertical bars, colored by severity) ---
site_bar_colors = site_summary["Severity"].map(ACTION_COLORS)
fig_sites_diff = go.Figure()
fig_sites_diff.add_trace(go.Bar(
    x=site_summary["Site"], y=site_summary["Diff %"] * 100,
    marker_color=site_bar_colors, name="Diff %",
))
fig_sites_diff.update_layout(
    plot_bgcolor=COLOR_PANEL, paper_bgcolor=COLOR_BG, font_color=COLOR_TEXT,
    height=220, margin=dict(l=10, r=10, t=10, b=10),
    xaxis=dict(gridcolor=COLOR_BORDER, tickangle=-35), yaxis=dict(gridcolor=COLOR_BORDER, title="Diff %"),
    showlegend=False,
)
st.plotly_chart(fig_sites_diff, use_container_width=True)

site_options = ["All sites"] + site_summary["Site"].tolist()
selected_site = st.selectbox("Filter by site", site_options)


def style_site_row(row):
    color = ACTION_COLORS.get(row["Severity"], COLOR_TEXT)
    styles = []
    for col in row.index:
        if col == "Total_GPS":
            styles.append(f"color: {COLOR_GPS};")
        elif col == "Total_MIS":
            styles.append(f"color: {COLOR_MIS};")
        elif col == "Diff %":
            styles.append(f"color: {color}; font-weight: 600;")
        else:
            styles.append("")
    return styles


site_styled = (
    site_summary.style
    .apply(style_site_row, axis=1)
    .format({"Total_GPS": "{:,.0f}", "Total_MIS": "{:,.0f}", "Diff %": "{:.1%}"})
    .hide(axis="columns", subset=["Severity"])
)
st.dataframe(site_styled, use_container_width=True, hide_index=True)

st.markdown("### Corrective actions")
st.caption("Click a card to filter the vehicle table below to just those vehicles.")
action_counts = df.groupby(["Action", "Severity"]).size().reset_index(name="Count")
order = ["Check GPS device", "File missing MIS log", "No data either side", "Audit MIS entries",
         "Verify unrecorded trips", "Keep an eye on it", "No action needed"]
action_counts["order"] = action_counts["Action"].apply(lambda a: order.index(a) if a in order else 99)
action_counts = action_counts.sort_values("order")

if "action_filter" not in st.session_state:
    st.session_state.action_filter = "All"

SEVERITY_EMOJI = {"critical": "🔴", "warn": "🟠", "watch": "🟡", "ok": "🟢"}

with st.container(key="corrective_actions"):
    cols = st.columns(len(action_counts)) if len(action_counts) else [st]
    for idx, (col, (_, row)) in enumerate(zip(cols, action_counts.iterrows())):
        is_active = st.session_state.action_filter == row["Action"]
        emoji = SEVERITY_EMOJI.get(row["Severity"], "⚪")
        label = f"{emoji} {row['Count']}\n{row['Action']}"
        if is_active:
            label = f"✓ {emoji} {row['Count']}\n{row['Action']}"
        with col:
            with st.container(key=f"actsev_{row['Severity']}_{idx}"):
                if st.button(label, key=f"actbtn_{row['Action']}", use_container_width=True):
                    st.session_state.action_filter = "All" if is_active else row["Action"]
                    st.rerun()

st.markdown("### Vehicles")
if st.session_state.action_filter != "All":
    fc1, fc2 = st.columns([5, 1])
    fc1.info(f"Filtered to action: **{st.session_state.action_filter}**")
    if fc2.button("✕ Clear filter", use_container_width=True):
        st.session_state.action_filter = "All"
        st.rerun()

col_a, col_b, col_c = st.columns([2, 1, 1])
search = col_a.text_input("Search vehicle no. or site")
only_flagged = col_b.checkbox("Flagged only")
action_filter = col_c.selectbox("Action", ["All"] + order, key="action_filter")

view = df.copy()
if selected_site != "All sites":
    view = view[view["Site"] == selected_site]
if only_flagged:
    view = view[view["Action"] != "No action needed"]
if action_filter != "All":
    view = view[view["Action"] == action_filter]
if search.strip():
    q = search.strip().lower()
    view = view[view["Vehicle"].str.lower().str.contains(q) | view["Site"].str.lower().str.contains(q)]

view = view.sort_values("Diff %", key=abs, ascending=False)
st.caption(f"{len(view)} vehicles")

display_cols = ["Vehicle", "Site", "Total GPS", "Total MIS", "Diff %", "Action", "Source", "Severity"]
display_df = view[display_cols]

ACTION_BG = {
    "critical": "rgba(255,93,93,0.16)",
    "warn": "rgba(245,166,35,0.16)",
    "watch": "rgba(245,166,35,0.10)",
    "ok": "rgba(61,220,132,0.14)",
}


def style_row(row):
    color = ACTION_COLORS.get(row["Severity"], COLOR_TEXT)
    bg = ACTION_BG.get(row["Severity"], "")
    styles = []
    for col in row.index:
        if col == "Total GPS":
            styles.append(f"color: {COLOR_GPS};")
        elif col == "Total MIS":
            styles.append(f"color: {COLOR_MIS};")
        elif col == "Diff %":
            styles.append(f"color: {color}; font-weight: 600;")
        elif col == "Action":
            styles.append(f"color: {color}; background-color: {bg}; font-weight: 500; border-radius: 5px;")
        else:
            styles.append("")
    return styles


styled = (
    display_df.style
    .apply(style_row, axis=1)
    .format({"Total GPS": "{:,.1f}", "Total MIS": "{:,.1f}", "Diff %": "{:.1%}"})
    .hide(axis="columns", subset=["Severity"])
)

st.dataframe(styled, use_container_width=True, hide_index=True, height=380)

st.markdown("### Vehicle drill-down")
veh_pick = st.selectbox("Pick a vehicle to see its daily GPS vs MIS chart", view["Vehicle"].tolist() if len(view) else df["Vehicle"].tolist())
if veh_pick:
    row = df[df["Vehicle"] == veh_pick].iloc[0]
    color = ACTION_COLORS[row["Severity"]]
    st.markdown(
        f"<div class='action-card' style='border-color:{color}55;'>"
        f"<b style='color:{color};'>{row['Action']}</b><br>"
        f"<span style='color:{COLOR_MUTED}; font-size:13px;'>{row['ActionDetail']}</span></div>",
        unsafe_allow_html=True,
    )
    vd = daily_df[daily_df["Vehicle"] == veh_pick]
    if has_daily and len(vd):
        fig2 = go.Figure()
        fig2.add_trace(go.Scatter(
            x=vd["Day"], y=vd["GPS"], name="GPS", mode="lines+markers",
            line=dict(color=COLOR_GPS, width=2.5), marker=dict(size=5),
            hovertemplate="Day %{x}<br>GPS: %{y:,.1f} km<extra></extra>",
        ))
        fig2.add_trace(go.Scatter(
            x=vd["Day"], y=vd["MIS"], name="MIS", mode="lines+markers",
            line=dict(color=COLOR_MIS, width=2.5), marker=dict(size=5),
            hovertemplate="Day %{x}<br>MIS: %{y:,.1f} km<extra></extra>",
        ))
        fig2.update_layout(
            plot_bgcolor=COLOR_PANEL, paper_bgcolor=COLOR_BG, font_color=COLOR_TEXT,
            height=260, margin=dict(l=10, r=10, t=10, b=10),
            xaxis=dict(gridcolor=COLOR_BORDER, title="Day of month", tickfont=dict(size=11), dtick=2),
            yaxis=dict(gridcolor=COLOR_BORDER, title="Km", tickfont=dict(size=11)),
            legend=dict(orientation="h", y=1.12, font=dict(size=12)),
            hovermode="x unified",
        )
        st.plotly_chart(fig2, use_container_width=True)
    else:
        st.caption("Daily chart isn't available for saved history months — only totals are stored.")

st.caption(f"Generated {datetime.now().strftime('%d %b %Y, %H:%M')} - runs entirely on your machine, nothing uploaded externally.")

# ---------------------------------------------------------------------------
# Monthly history — trend across all saved months + manage saved entries
# ---------------------------------------------------------------------------
hist = load_history()  # reload in case this run just saved a new entry
if hist:
    st.markdown("---")
    st.markdown("### Monthly history")
    st.caption("Every file you upload is saved here automatically under its month label, even after you upload a different file or restart the app.")

    sorted_entries = sorted(hist.values(), key=lambda x: x["key"])
    hist_df = pd.DataFrame([{
        "Month": e["label"], "Total GPS": e["total_gps"], "Total MIS": e["total_mis"],
        "Diff %": ((e["total_mis"] - e["total_gps"]) / e["total_gps"]) if e["total_gps"] else 0,
        "Vehicles": e["vehicles"], "Flagged": e.get("flagged", "-"),
    } for e in sorted_entries])

    fig3 = go.Figure()
    fig3.add_trace(go.Scatter(x=hist_df["Month"], y=hist_df["Total GPS"], name="GPS", line=dict(color=COLOR_GPS, width=2), mode="lines+markers"))
    fig3.add_trace(go.Scatter(x=hist_df["Month"], y=hist_df["Total MIS"], name="MIS", line=dict(color=COLOR_MIS, width=2), mode="lines+markers"))
    fig3.update_layout(
        plot_bgcolor=COLOR_PANEL, paper_bgcolor=COLOR_BG, font_color=COLOR_TEXT,
        height=260, margin=dict(l=10, r=10, t=10, b=10),
        xaxis=dict(gridcolor=COLOR_BORDER), yaxis=dict(gridcolor=COLOR_BORDER),
        legend=dict(orientation="h", y=1.15),
    )
    st.plotly_chart(fig3, use_container_width=True)

    st.dataframe(
        hist_df.style.format({"Total GPS": "{:,.0f}", "Total MIS": "{:,.0f}", "Diff %": "{:.1%}"}),
        use_container_width=True, hide_index=True,
    )

    with st.expander("Remove a saved month"):
        del_choice = st.selectbox("Month to remove", [e["label"] for e in sorted_entries])
        if st.button("Delete this month from history"):
            key_to_delete = next(e["key"] for e in sorted_entries if e["label"] == del_choice)
            del hist[key_to_delete]
            save_history(hist)
            st.success(f"Removed {del_choice}. Refresh the page to see the updated list.")
